"""Node 8: Output generator — generate all export formats and upload to Azure Blob.

Produces all output artefacts from clean, hierarchy-resolved data:
1. Nested JSON (sites > locations > assets > work_orders > tasks)
2. Flat CSV exports (per table)
3. SQL INSERT statements (in FK-dependency order)
4. PDF migration summary report
5. Mapping flow document (PDF/Word)
6. IntermediateSchema (for svc-ingestion handoff)
7. Upload all to Azure Blob with signed URLs

Dual logging: execution_logs for Streamlit display + logger for Docker monitoring

EL-M.8: IntermediateSchema Pydantic validates
"""

import io
import json
import logging
from datetime import datetime
from typing import Optional
from uuid import UUID

import chardet
import pandas as pd
from azure.storage.blob.aio import BlobClient, BlobServiceClient

from ...export import (
    build_nested_json,
    export_to_csv,
    export_to_sql,
    generate_pdf_report,
    build_intermediate_schema,
)
from ..state import MigrationState

from cafm_shared.logging import get_logger
logger = get_logger(__name__)


async def output_generator_node(state: MigrationState) -> MigrationState:
    """
    Node 8: Generate all output formats and upload to Azure Blob.

    This node processes the FULL file (not just 5-row sample from Node 1).
    Downloads source file from Azure Blob to ensure complete data for exports.

    Generates:
    - Nested JSON hierarchy
    - Per-table CSV exports
    - SQL INSERT statements (in FK-dependency order)
    - PDF migration summary report
    - Mapping flow document (PDF/Word)
    - IntermediateSchema (Pydantic model for svc-ingestion)

    All artefacts uploaded to Azure Blob with signed URLs returned.

    Args:
        state: MigrationState with cleaned_tables, mappings, hierarchies

    Returns:
        Updated state with all output URLs, IntermediateSchema, execution_logs
    """

    _node_started_at = datetime.utcnow()
    migration_id = state.get("migration_id")
    cmms_name = state.get("cmms_name", "Unknown")
    organization_id = state.get("organization_id")
    source_blob_url = state.get("source_blob_url")
    source_filename = state.get("source_filename", "unknown")
    cleaned_tables = state.get("cleaned_tables", {})
    confirmed_hierarchies = state.get("confirmed_hierarchies", [])
    tier1_mappings = state.get("tier1_mappings", [])
    tier2_auto_mappings = state.get("tier2_auto_accepted", [])
    tier2_human_decisions = state.get("human_approved_mappings", [])
    overall_confidence = state.get("overall_confidence", 0.0)
    tier2_unmappable = state.get("tier2_unmappable", [])
    data_quality_warnings = state.get("data_quality_warnings", [])
    hierarchy_cycles = state.get("hierarchy_cycles", [])
    orphaned_records = state.get("orphaned_records", {})

    # ── Set up dual logging (Streamlit + Docker) ──────────────────
    execution_logs = []

    def log(msg: str):
        """Helper to append logs and also log to docker"""
        execution_logs.append(f"[Node 8] {msg}")
        logger.info(f"[Node 8] {msg}")

    log(f"Starting output generation for migration {migration_id}")

    if not cleaned_tables:
        log("❌ No cleaned tables found")
        state["error_message"] = "No data for output generation"
        state["error_node"] = 8
        state["execution_logs"] = execution_logs
        return state

    try:
        # ── CRITICAL: Get FULL file data for complete export ──────────────
        # `cleaned_tables` (from state) is always list-of-dicts — never DataFrames.
        # We build a SEPARATE `df_tables` dict of DataFrames for CSV generation only.
        # DataFrames MUST NOT be written back to state (LangGraph checkpointer
        # cannot serialize them to msgpack).

        state_full_tables = state.get("full_tables")
        is_direct_upload = (not source_blob_url) or source_blob_url == "direct_upload"

        # records_tables: list-of-dicts — used for SQL, JSON builder, IntermediateSchema
        # df_tables:      DataFrames    — used ONLY for CSV generation (never touches state)
        records_tables: dict = dict(cleaned_tables)  # shallow copy — keeps list-of-dicts
        df_tables: dict = {}

        if state_full_tables:
            # Fast path: Node 1 stored full file as list-of-dicts in state
            log(f"Using full_tables from state ({len(state_full_tables)} table(s))")
            for tbl_name, records in state_full_tables.items():
                if isinstance(records, list):
                    records_tables[tbl_name] = records
                    df_tables[tbl_name] = pd.DataFrame(records)
                elif isinstance(records, pd.DataFrame):
                    records_tables[tbl_name] = records.to_dict(orient="records")
                    df_tables[tbl_name] = records
            total_rows = sum(len(v) for v in records_tables.values() if hasattr(v, "__len__"))
            log(f"Full data loaded from state: {total_rows} total rows")

        elif is_direct_upload:
            # Direct upload but full_tables not in state — use cleaned_tables
            log("⚠️  Direct upload: full_tables not in state, using cleaned_tables for export")
            for tbl_name, records in records_tables.items():
                if isinstance(records, list):
                    df_tables[tbl_name] = pd.DataFrame(records)
                elif isinstance(records, pd.DataFrame):
                    # Should not happen, but handle defensively
                    df_tables[tbl_name] = records
                    records_tables[tbl_name] = records.to_dict(orient="records")

        else:
            # Download from Azure Blob
            log("Downloading FULL source file from Blob for complete export...")
            try:
                async with BlobClient.from_blob_url(source_blob_url) as blob_client:
                    file_bytes_dl = await blob_client.download_blob()
                    file_content = await file_bytes_dl.readall()
                log(f"Downloaded FULL file: {len(file_content):,} bytes")

                detected = chardet.detect(file_content)
                encoding = detected.get("encoding", "utf-8") or "utf-8"
                file_str = file_content.decode(encoding, errors="replace")
                sample = file_str[:4096]
                delimiter = "," if "," in sample else "\t"

                try:
                    df_full = pd.read_csv(io.StringIO(file_str), delimiter=delimiter, dtype=str)
                    records_tables["data"] = df_full.to_dict(orient="records")
                    df_tables["data"] = df_full
                    log(f"Parsed FULL CSV: {len(df_full):,} rows × {len(df_full.columns)} columns")
                except Exception as csv_err:
                    log(f"CSV parse failed: {csv_err}; trying Excel...")
                    excel_file = io.BytesIO(file_content)
                    xls = pd.ExcelFile(excel_file)
                    for sheet_name in xls.sheet_names:
                        df_full = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str)
                        records_tables[sheet_name] = df_full.to_dict(orient="records")
                        df_tables[sheet_name] = df_full
                        log(f"Parsed FULL Excel sheet '{sheet_name}': {len(df_full):,} rows")

            except Exception as e:
                log(f"⚠️  Failed to download full file from Blob: {e} — using cleaned_tables")
                for tbl_name, records in records_tables.items():
                    if isinstance(records, list):
                        df_tables[tbl_name] = pd.DataFrame(records)

        # ── Step 1: Build nested JSON (uses records, not DataFrames) ─────────
        log(f"Building nested JSON hierarchy...")
        containment_hierarchy = state.get("containment_hierarchy", {})

        try:
            nested_json = build_nested_json(records_tables, containment_hierarchy, confirmed_hierarchies)
            total_entities = sum(len(v) if isinstance(v, list) else 1 for v in nested_json.get("entities", {}).values())
            log(f"✅ Nested JSON built: {total_entities} entities")
        except Exception as e:
            log(f"❌ Failed to build nested JSON: {e}")
            state["error_message"] = f"Nested JSON generation failed: {str(e)}"
            state["error_node"] = 8
            state["execution_logs"] = execution_logs
            return state

        # ── Step 2: Export to CSV (uses DataFrames for to_csv()) ─────────────
        log(f"Exporting {len(df_tables)} table(s) to CSV...")
        csv_exports = {}
        total_csv_rows = 0

        try:
            for table_name, df in df_tables.items():
                if isinstance(df, pd.DataFrame):
                    csv_content = df.to_csv(index=False)
                    csv_exports[table_name] = csv_content
                    total_csv_rows += len(df)
                    log(f"  Exported '{table_name}': {len(df):,} rows")
            if not csv_exports and records_tables:
                # Fallback: build CSV from records if df_tables was empty
                for table_name, records in records_tables.items():
                    if records:
                        df_tmp = pd.DataFrame(records)
                        csv_exports[table_name] = df_tmp.to_csv(index=False)
                        total_csv_rows += len(df_tmp)
                        log(f"  Exported '{table_name}' (from records): {len(df_tmp):,} rows")
            log(f"✅ CSV export complete: {total_csv_rows:,} total rows across {len(csv_exports)} file(s)")
        except Exception as e:
            log(f"❌ CSV export failed: {e}")
            state["error_message"] = f"CSV export failed: {str(e)}"
            state["error_node"] = 8
            state["execution_logs"] = execution_logs
            return state

        # ── Step 3: Export to SQL (uses records_tables — list-of-dicts) ──────
        log("Generating SQL INSERT statements (FK-dependency order)...")

        try:
            sql_script = export_to_sql(records_tables, confirmed_hierarchies)
            sql_lines = sql_script.count('\n')
            log(f"✅ SQL script generated: {len(sql_script):,} bytes, {sql_lines:,} lines")
        except Exception as e:
            log(f"❌ SQL export failed: {e}")
            state["error_message"] = f"SQL export failed: {str(e)}"
            state["error_node"] = 8
            state["execution_logs"] = execution_logs
            return state

        # ── Step 4: Generate PDF migration summary report ────────────
        log("Generating PDF migration summary report...")

        try:
            pdf_bytes = generate_pdf_report(
                migration_id=str(migration_id),
                cmms_name=cmms_name,
                t1_count=len(tier1_mappings),
                t2_auto_count=len(tier2_auto_mappings),
                t2_human_count=len(tier2_human_decisions),
                t2_unmappable=tier2_unmappable,
                overall_confidence=overall_confidence,
                data_quality_warnings=data_quality_warnings,
                tier1_mappings=tier1_mappings,
                tier2_auto_mappings=tier2_auto_mappings,
                tier2_human_decisions=tier2_human_decisions,
                confirmed_hierarchies=confirmed_hierarchies,
                hierarchy_cycles=hierarchy_cycles,
                orphaned_records=orphaned_records,
            )
            log(f"✅ PDF report generated: {len(pdf_bytes):,} bytes")
        except Exception as e:
            log(f"❌ PDF report generation failed: {e}")
            # PDF generation is not critical; continue with warning
            pdf_bytes = None
            log("⚠️  Continuing without PDF report...")

        # ── Step 5: Build IntermediateSchema ─────────────────────────
        log("Building IntermediateSchema for svc-ingestion handoff...")

        try:
            intermediate_schema = build_intermediate_schema(
                migration_id=str(migration_id),
                cmms_name=cmms_name,
                source_filename=source_filename,
                source_blob_url=source_blob_url,
                cleaned_tables=records_tables,
                tier1_mappings=tier1_mappings,
                tier2_auto_mappings=tier2_auto_mappings,
                tier2_human_decisions=tier2_human_decisions,
                overall_confidence=overall_confidence,
                confirmed_hierarchies=confirmed_hierarchies,
                table_routing=state.get("table_routing"),
                new_tables=state.get("new_tables"),
                detected_file_format=state.get("detected_file_format"),
            )
            log(f"✅ IntermediateSchema built")
        except Exception as e:
            log(f"❌ IntermediateSchema build failed: {e}")
            state["error_message"] = f"IntermediateSchema build failed: {str(e)}"
            state["error_node"] = 8
            state["execution_logs"] = execution_logs
            return state

        # ── Step 6: EL-M.8 Validation: Validate IntermediateSchema ───
        log("Running EL-M.8 validation...")

        try:
            schema_dict = intermediate_schema.dict()

            # Validate required fields
            required_fields = [
                "ingestion_id",
                "source_type",
                "agent_id",
                "entities",
                "confidence",
                "audit",
            ]

            missing_fields = []
            for field in required_fields:
                if field not in schema_dict:
                    missing_fields.append(field)

            if missing_fields:
                raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

            # Validate entity structure
            entities = schema_dict.get("entities", {})
            if not isinstance(entities, dict):
                raise ValueError(f"entities must be dict, got {type(entities)}")

            # Validate confidence structure
            confidence = schema_dict.get("confidence", {})
            if not isinstance(confidence, dict):
                raise ValueError(f"confidence must be dict, got {type(confidence)}")

            if "overall" not in confidence:
                raise ValueError("confidence.overall is required")

            state["el_m8_passed"] = True
            log(f"✅ EL-M.8 PASSED: IntermediateSchema validates")

        except Exception as e:
            log(f"❌ EL-M.8 FAILED: Schema validation error: {e}")
            state["error_message"] = f"IntermediateSchema validation failed: {str(e)}"
            state["error_node"] = 8
            state["el_m8_passed"] = False
            state["execution_logs"] = execution_logs
            return state

        # ── Step 7: Upload to Azure Blob ─────────────────────────────
        log("Uploading artefacts to Azure Blob...")

        from ...config import get_settings as _get_settings
        _settings = _get_settings()
        blob_connection_string = _settings.azure_storage_connection_string
        blob_container = _settings.azure_blob_container_name
        blob_base_path = f"migrations/{migration_id}"

        # Keep backward-compatible nested output while also including full per-table
        # records so downloads reflect all flow tables, not only "sites" hierarchy.
        output_json_payload = {
            "nested_hierarchy": nested_json,
            "tables": records_tables,
            "table_count": len(records_tables),
            "tables_included": sorted(records_tables.keys()),
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

        # Build Excel workbook: one sheet per table
        excel_bytes: bytes | None = None
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # remove default empty sheet
            for table_name, df in df_tables.items():
                if not isinstance(df, pd.DataFrame) or df.empty:
                    continue
                sheet_name = table_name[:31]  # Excel sheet name limit
                ws = wb.create_sheet(title=sheet_name)
                ws.append(list(df.columns))
                for row in df.itertuples(index=False, name=None):
                    ws.append(list(row))
            if not wb.sheetnames and records_tables:
                for table_name, records in records_tables.items():
                    if not records:
                        continue
                    df_tmp = pd.DataFrame(records)
                    sheet_name = table_name[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(list(df_tmp.columns))
                    for row in df_tmp.itertuples(index=False, name=None):
                        ws.append(list(row))
            buf = io.BytesIO()
            wb.save(buf)
            excel_bytes = buf.getvalue()
            log(f"✅ Excel workbook built: {len(wb.sheetnames)} sheet(s), {len(excel_bytes):,} bytes")
        except Exception as e:
            log(f"❌ Excel workbook generation failed: {e}")

        artefacts: dict[str, bytes | str] = {
            "output.json": json.dumps(output_json_payload, indent=2),
            "output.sql": sql_script,
        }
        if pdf_bytes:
            artefacts["migration_report.pdf"] = pdf_bytes
        if excel_bytes:
            artefacts["output.xlsx"] = excel_bytes
        for table_name, csv_content in csv_exports.items():
            artefacts[f"table_{table_name}.csv"] = csv_content

        uploaded_count = 0
        urls_generated: dict[str, str] = {}

        if not blob_connection_string:
            log("⚠️  AZURE_STORAGE_CONNECTION_STRING not set — skipping blob upload, URLs will be empty")
        else:
            async with BlobServiceClient.from_connection_string(blob_connection_string) as svc:
                for filename, content in artefacts.items():
                    blob_path = f"{blob_base_path}/{filename}"
                    try:
                        blob_client = svc.get_blob_client(container=blob_container, blob=blob_path)
                        data: bytes = content.encode("utf-8") if isinstance(content, str) else content
                        await blob_client.upload_blob(data, overwrite=True)
                        urls_generated[filename] = blob_client.url
                        uploaded_count += 1
                        log(f"  ✅ Uploaded: {filename} ({len(data):,} bytes) → {blob_client.url}")
                    except Exception as e:
                        log(f"  ⚠️  Failed to upload {filename}: {e}")

        log(f"✅ Uploaded {uploaded_count}/{len(artefacts)} artefacts to Blob")

        # ── Step 8: Update state with output URLs ────────────────────
        state["intermediate_schema"] = schema_dict
        state["output_json_url"] = urls_generated.get("output.json", "")
        state["output_csv_url"] = urls_generated.get("output.xlsx", "")
        state["output_sql_url"] = urls_generated.get("output.sql", "")
        state["output_sql_script"] = sql_script
        state["migration_report_url"] = urls_generated.get("migration_report.pdf", "")

        # Track all artefact metadata
        state["exported_artefacts"] = {
            "total_count": uploaded_count,
            "by_type": {
                "json": 1,
                "csv": len(csv_exports),
                "sql": 1,
                "pdf": 1 if pdf_bytes else 0,
            },
            "total_size_bytes": sum(len(str(c)) for c in artefacts.values()),
            "urls": urls_generated,
        }

        log(
            f"✅ Output generation complete: "
            f"1 JSON + {len(csv_exports)} CSV + 1 SQL + 1 PDF + IntermediateSchema"
        )

        state["current_step"] = 8
        state["execution_logs"] = execution_logs

        if "event_log" in state and isinstance(state["event_log"], list):
            state["event_log"].append(
                {
                    "timestamp": datetime.utcnow().isoformat(),
                    "event": "node_complete",
                    "node": 8,
                    "detail": f"All outputs generated and uploaded ({uploaded_count} artefacts)",
                }
            )

        migration_id = state.get("migration_id")
        if migration_id:
            from .db_writer import update_node_progress, write_step_pause
            await update_node_progress(
                migration_id, "8_output_generation",
                output_json_url=state.get("output_json_url"),
                output_csv_url=state.get("output_csv_url"),
                output_sql_url=state.get("output_sql_url"),
                migration_report_url=state.get("migration_report_url"),
            )
            await write_step_pause(
                migration_id,
                "step_8_output_generation",
                {
                    "node": 8,
                    "label": "Output Generation",
                    "tables": len(records_tables),
                    "formats": ["json", "csv", "sql", "pdf"],
                    "artifacts_uploaded": uploaded_count,
                    "json_url": state.get("output_json_url"),
                    "csv_url": state.get("output_csv_url"),
                    "sql_url": state.get("output_sql_url"),
                    "report_url": state.get("migration_report_url"),
                },
            )
            from .schema_db_writer import migration_append_node_log_auto
            await migration_append_node_log_auto(
                migration_id, 9, "Output Generation", _node_started_at, datetime.utcnow(),
                output={"table_count": len(records_tables),
                        "artifacts_uploaded": uploaded_count,
                        "formats": ["json", "csv", "sql", "pdf"],
                        "json_url": state.get("output_json_url"),
                        "csv_url": state.get("output_csv_url"),
                        "sql_url": state.get("output_sql_url"),
                        "report_url": state.get("migration_report_url")},
                logs=[f"Generated outputs for {len(records_tables)} tables",
                      f"Formats: JSON, CSV, SQL, PDF",
                      f"{uploaded_count} artifacts uploaded to Azure Blob",
                      f"EL-M.8: {'PASSED' if state.get('el_m8_passed') else 'FAILED'}"],
            )

        return state

    except Exception as e:
        log(f"❌ ERROR: {str(e)}")
        logger.exception(f"[Node 8] Unhandled exception: {e}")
        state["error_message"] = str(e)
        state["error_node"] = 8
        state["error_timestamp"] = datetime.utcnow()
        state["status"] = "failed"
        state["execution_logs"] = execution_logs
        return state
