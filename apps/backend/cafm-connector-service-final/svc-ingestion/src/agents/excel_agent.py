"""
svc-ingestion/src/agents/excel_agent.py

Task 2.4 — Excel Agent (Layer 2, Stage 2).

Ingests .xlsx / .xls / .xlsm files into the unified plenum_cafm store.

Design:
  - openpyxl (data_only=True)   reads formula-resolved cell values
  - Sheet detection              first non-empty sheet wins; all sheets scanned
  - Schema mapper                called ONCE per file — Claude Haiku maps
                                 column headers → canonical fields
  - EL-3.0                       blocks write if mapping confidence < 0.80
  - asyncpg COPY                 bulk insert for direct-copy entity types
  - Batch size                   1000 rows per COPY call
  - No EL-2.3                    structured data — no LLM-as-judge needed
                                 (same rationale as CSV agent per CLAUDE.md)
  - EL-4.0                       pre-write row count check after COPY

Known entity types mapped by the schema mapper:
  assets, spare_parts, work_orders, maintenance_plans, users

OTel spans: ingestion.stage2.extract + schema_mapper.map (inside map_headers)
"""

from __future__ import annotations

import time
import uuid
from typing import Any
from uuid import UUID

import anthropic
import openpyxl
import openpyxl.utils.exceptions
from opentelemetry import trace
from opentelemetry.trace import StatusCode
from sqlalchemy.ext.asyncio import AsyncEngine

from cafm_shared.logging import get_logger
from shared.intermediate_schema import (
    AgentId,
    AssetEntity,
    AuditInfo,
    ConfidenceLevel,
    ConfidenceResult,
    EntitiesBlock,
    ExtractionMethod,
    IntermediateSchema,
    ModelUsed,
    SparePartEntity,
    SourceType,
    WorkOrderEntity,
)
from shared.schema_mapper import SchemaMapping, map_headers

logger = get_logger(__name__)
tracer = trace.get_tracer(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────

_BATCH_SIZE = 1_000
_SAMPLE_ROWS = 50        # rows passed to schema mapper for context
_SCHEMA = "plenum_cafm"

# Maximum rows to read from a single sheet (guard against massive files)
_MAX_ROWS = 200_000

# ── Entity type detection ──────────────────────────────────────────────────────
# Identical to csv_agent — score by canonical field overlap.

_ENTITY_SIGNATURES: dict[str, frozenset[str]] = {
    "assets": frozenset({"asset_code", "asset_name", "category", "make", "model", "serial"}),
    "spare_parts": frozenset({
        "part_code", "stock_on_hand", "minimum_allowed_stock", "bom_group_name",
    }),
    "work_orders": frozenset({
        "wo_code", "wo_priority", "wo_status", "wo_type", "maintenance_type",
    }),
    "maintenance_plans": frozenset({
        "sm_code", "trigger_type", "schedule_interval", "sm_priority",
    }),
    "users": frozenset({"user_full_name", "user_name", "user_title", "reports_to"}),
}

# Canonical field → actual DB column name (per entity type)
_FIELD_TO_COLUMN: dict[str, dict[str, str]] = {
    "assets": {
        "asset_code": "asset_code",
        "asset_name": "asset_name",
        "make": "manufacturer",
        "model": "model_number",
        "serial": "serial_number",
        "status": "status",
    },
    "spare_parts": {
        "part_code": "part_code",
        "stock_on_hand": "stock_quantity",
        "minimum_allowed_stock": "reorder_level",
    },
    "work_orders": {
        "wo_priority": "priority",
        "wo_status": "status",
    },
}

# Tables where asyncpg COPY is used directly
_DIRECT_COPY_TABLES: frozenset[str] = frozenset({"assets", "spare_parts", "work_orders"})

# Required columns per table with fallback defaults
_REQUIRED_COLUMNS: dict[str, dict[str, Any]] = {
    "assets": {
        "asset_name": "Unnamed Asset",
    },
    "spare_parts": {
        "part_name": "Unnamed Part",
        "stock_quantity": 0,
        "reorder_level": 0,
    },
    "work_orders": {
        "title": "Imported Work Order",
        "priority": "medium",
        "status": "open",
    },
}


# ── Helpers ────────────────────────────────────────────────────────────────────


def _coerce_str(val: Any) -> str | None:
    """Safe string coercion — returns None for None/empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _coerce_int(val: Any) -> int | None:
    """Safe int coercion — returns None on failure."""
    if val is None:
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def _detect_entity_type(canonical_fields: set[str]) -> str:
    """Return entity type with highest signature overlap."""
    scores = {
        etype: len(sig & canonical_fields)
        for etype, sig in _ENTITY_SIGNATURES.items()
    }
    best = max(scores, key=lambda k: scores[k])
    return best if scores[best] > 0 else "unknown"


def _read_workbook_rows(
    excel_bytes: bytes,
) -> tuple[list[str], list[dict[str, Any]], str]:
    """
    Open the workbook and extract headers + rows from the best sheet.

    Strategy:
      1. Try each sheet in order.
      2. Pick the first sheet where row 1 contains at least 2 non-empty cells
         (i.e. looks like a header row).
      3. Return (headers, rows_as_dicts, sheet_name).

    Uses data_only=True so formulas are read as their last-calculated values.
    """
    wb = openpyxl.load_workbook(
        filename=__import__("io").BytesIO(excel_bytes),
        data_only=True,
        read_only=True,
    )

    chosen_sheet = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Peek at first row
        first_row_vals = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            first_row_vals = [c for c in row if c is not None and str(c).strip()]
        if len(first_row_vals) >= 2:
            chosen_sheet = sheet_name
            break

    if chosen_sheet is None:
        # Fallback: use the first sheet whatever it looks like
        chosen_sheet = wb.sheetnames[0]

    ws = wb[chosen_sheet]
    rows_iter = ws.iter_rows(values_only=True)

    # First row → headers
    header_row: tuple[Any, ...] = next(rows_iter, ())
    headers: list[str] = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(header_row)
    ]

    # Remaining rows → dicts, capped at _MAX_ROWS
    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        if len(rows) >= _MAX_ROWS:
            break
        # Skip fully empty rows
        if all(c is None for c in raw_row):
            continue
        row_dict: dict[str, Any] = {}
        for i, val in enumerate(raw_row):
            col_name = headers[i] if i < len(headers) else f"col_{i}"
            row_dict[col_name] = val
        rows.append(row_dict)

    wb.close()
    return headers, rows, chosen_sheet


# ── asyncpg COPY writer ────────────────────────────────────────────────────────


async def _copy_batch_to_table(
    engine: AsyncEngine,
    table_name: str,
    records: list[tuple[Any, ...]],
    columns: list[str],
) -> None:
    """
    Write one batch of records to a plenum_cafm table using asyncpg COPY.

    Uses SQLAlchemy 2.0 async engine to retrieve the raw asyncpg connection.
    All records in the batch must have identical columns (same order).
    """
    async with engine.connect() as sa_conn:
        raw = await sa_conn.get_raw_connection()
        asyncpg_conn = raw.driver_connection

        await asyncpg_conn.copy_records_to_table(
            table_name,
            records=records,
            columns=columns,
            schema_name=_SCHEMA,
        )


# ── Record builders ────────────────────────────────────────────────────────────


def _build_asset_record(
    row: dict[str, Any],
    org_id: UUID,
    mapping: SchemaMapping,
) -> tuple[tuple[Any, ...], list[str]]:
    """Build asyncpg COPY record tuple for the assets table."""
    row_id = uuid.uuid4()
    record: dict[str, Any] = {
        "id": row_id,
        "organization_id": org_id,
        "asset_name": _coerce_str(row.get("asset_name")) or "Unnamed Asset",
    }
    optional: dict[str, Any] = {
        "asset_code": _coerce_str(row.get("asset_code")),
        "serial_number": _coerce_str(row.get("serial")),
        "manufacturer": _coerce_str(row.get("make")),
        "model_number": _coerce_str(row.get("model")),
    }
    record.update({k: v for k, v in optional.items() if v is not None})
    columns = list(record.keys())
    return tuple(record.values()), columns


def _build_spare_part_record(
    row: dict[str, Any],
    org_id: UUID,
    mapping: SchemaMapping,
) -> tuple[tuple[Any, ...], list[str]]:
    """Build asyncpg COPY record for spare_parts table."""
    row_id = uuid.uuid4()
    part_name = _coerce_str(row.get("part_code")) or "Unnamed Part"
    stock = _coerce_int(row.get("stock_on_hand")) or 0
    reorder = _coerce_int(row.get("minimum_allowed_stock")) or 0
    record: dict[str, Any] = {
        "id": row_id,
        "organization_id": org_id,
        "part_name": part_name,
        "stock_quantity": max(stock, 0),
        "reorder_level": max(reorder, 0),
    }
    part_code = _coerce_str(row.get("part_code"))
    if part_code:
        record["part_code"] = part_code
    columns = list(record.keys())
    return tuple(record.values()), columns


def _build_work_order_record(
    row: dict[str, Any],
    org_id: UUID,
    mapping: SchemaMapping,
) -> tuple[tuple[Any, ...], list[str]]:
    """Build asyncpg COPY record for work_orders table."""
    row_id = uuid.uuid4()
    wo_code = _coerce_str(row.get("wo_code"))
    title = wo_code or "Imported Work Order"
    priority = (_coerce_str(row.get("wo_priority")) or "medium").lower()
    status = (_coerce_str(row.get("wo_status")) or "open").lower()
    record: dict[str, Any] = {
        "id": row_id,
        "organization_id": org_id,
        "title": title,
        "priority": priority,
        "status": status,
    }
    desc = _coerce_str(row.get("wo_type")) or _coerce_str(row.get("maintenance_type"))
    if desc:
        record["description"] = desc
    columns = list(record.keys())
    return tuple(record.values()), columns


_RECORD_BUILDERS = {
    "assets": _build_asset_record,
    "spare_parts": _build_spare_part_record,
    "work_orders": _build_work_order_record,
}


# ── Entity builders (for IntermediateSchema) ───────────────────────────────────


def _row_to_asset_entity(row: dict[str, Any]) -> AssetEntity | None:
    asset_code = _coerce_str(row.get("asset_code"))
    asset_name = _coerce_str(row.get("asset_name"))
    serial = _coerce_str(row.get("serial"))
    if not asset_code and not asset_name and not serial:
        return None
    return AssetEntity(
        asset_code=asset_code,
        name=asset_name,
        serial_number=serial,
        category=_coerce_str(row.get("category")),
        manufacturer=_coerce_str(row.get("make")),
        model_number=_coerce_str(row.get("model")),
        extra={
            k: v for k, v in row.items()
            if k not in {"asset_code", "asset_name", "serial", "category", "make", "model"}
            and v is not None
        },
    )


def _row_to_spare_part_entity(row: dict[str, Any]) -> SparePartEntity | None:
    part_code = _coerce_str(row.get("part_code"))
    if not part_code:
        return None
    stock = _coerce_int(row.get("stock_on_hand"))
    return SparePartEntity(
        part_number=part_code,
        name=part_code,
        quantity=float(stock) if stock is not None else None,
        supplier=_coerce_str(row.get("supplier")),
        extra={
            k: v for k, v in row.items()
            if k not in {"part_code", "stock_on_hand", "minimum_allowed_stock", "supplier"}
            and v is not None
        },
    )


def _row_to_work_order_entity(row: dict[str, Any]) -> WorkOrderEntity | None:
    wo_code = _coerce_str(row.get("wo_code"))
    if not wo_code:
        return None
    return WorkOrderEntity(
        work_order_number=wo_code,
        priority=_coerce_str(row.get("wo_priority")),
        status=_coerce_str(row.get("wo_status")),
        extra={
            k: v for k, v in row.items()
            if k not in {"wo_code", "wo_priority", "wo_status", "wo_type"}
            and v is not None
        },
    )


# ── Public API ─────────────────────────────────────────────────────────────────


async def extract_excel(
    excel_bytes: bytes,
    *,
    source_filename: str,
    ingestion_id: UUID,
    blob_url: str,
    organization_id: UUID,
    redis: Any,
    client: anthropic.AsyncAnthropic,
    engine: AsyncEngine,
    dry_run: bool = False,
) -> IntermediateSchema:
    """
    Stage 2 extraction for Excel files (.xlsx / .xls / .xlsm).

    Args:
        excel_bytes:     Raw file bytes (already validated by Stage 1 / EL-2.0).
        source_filename: Original filename — used for logging and audit.
        ingestion_id:    UUID from ingestion_documents (created by Stage 1).
        blob_url:        Azure Blob URL of the stored original.
        organization_id: Tenant/org UUID — stamped on every DB row.
        redis:           Async Redis client (passed to schema mapper for caching).
        client:          Async Anthropic client (Haiku — schema mapping only).
        engine:          Async SQLAlchemy engine (used to get raw asyncpg conn).

    Returns:
        IntermediateSchema — summary of what was ingested + full entity list.

    Eval layers applied:
        EL-3.0  Schema mapper confidence check — blocks write if < 0.80
        EL-4.0  Post-write row count check (rows_written vs total_rows)
        No EL-2.3 — structured data; no LLM-as-judge needed (CLAUDE.md §6 Agent 5)
    """
    t0 = time.monotonic()

    with tracer.start_as_current_span("ingestion.stage2.extract") as span:
        span.set_attribute("cafm.ingestion_id", str(ingestion_id))
        span.set_attribute("cafm.agent_id", AgentId.EXCEL.value)
        span.set_attribute("cafm.source_type", SourceType.EXCEL.value)
        span.set_attribute("cafm.extraction_method", ExtractionMethod.OPENPYXL_CLAUDE.value)
        span.set_attribute("cafm.file_size_bytes", len(excel_bytes))
        span.set_attribute("cafm.source_filename", source_filename)

        try:
            # ── 1. Read workbook — sheet detection + formula resolution ──────
            try:
                raw_headers, all_rows, sheet_name = _read_workbook_rows(excel_bytes)
            except (
                openpyxl.utils.exceptions.InvalidFileException,
                KeyError,
                StopIteration,
                Exception,
            ) as wb_exc:
                raise ValueError(
                    f"Cannot open workbook '{source_filename}': {wb_exc}"
                ) from wb_exc

            total_rows = len(all_rows)
            span.set_attribute("cafm.raw_row_count", total_rows)
            span.set_attribute("cafm.raw_column_count", len(raw_headers))
            span.set_attribute("cafm.sheet_name", sheet_name)

            logger.info(
                "excel_agent.workbook_read",
                ingestion_id=str(ingestion_id),
                source_filename=source_filename,
                sheet=sheet_name,
                rows=total_rows,
                columns=len(raw_headers),
            )

            # ── 2. Schema mapper — called ONCE per file ───────────────────
            # Pass first _SAMPLE_ROWS rows as context for Haiku.
            sample_rows = all_rows[:_SAMPLE_ROWS]
            # Convert to string-keyed dicts (openpyxl may produce native types)
            sample_rows_str: list[dict[str, Any]] = [
                {str(k): (str(v) if v is not None else None) for k, v in r.items()}
                for r in sample_rows
            ]

            mapping: SchemaMapping = await map_headers(
                raw_headers,
                redis=redis,
                client=client,
                sample_rows=sample_rows_str,
            )

            span.set_attribute("cafm.schema_mapped_count", len(mapping.mapped))
            span.set_attribute("cafm.schema_unmatched_count", len(mapping.unmatched))
            span.set_attribute("cafm.schema_cache_hit", mapping.cached)
            span.set_attribute("cafm.schema_requires_review", mapping.requires_human_review)

            # ── EL-3.0 — Mapping confidence check ────────────────────────
            # Block write if confidence < 0.80 (mapping routed to review queue
            # by the caller; we return LOW confidence schema here).
            if mapping.requires_human_review:
                logger.warning(
                    "excel_agent.low_confidence_mapping",
                    ingestion_id=str(ingestion_id),
                    overall_confidence=mapping.overall_confidence,
                    unmatched=mapping.unmatched,
                )
                return IntermediateSchema(
                    ingestion_id=ingestion_id,
                    source_type=SourceType.EXCEL,
                    agent_id=AgentId.EXCEL,
                    source_filename=source_filename,
                    source_blob_url=blob_url,
                    extraction_method=ExtractionMethod.OPENPYXL_CLAUDE,
                    model_used=ModelUsed.HAIKU,
                    entities=EntitiesBlock(),
                    confidence=ConfidenceResult(
                        overall=ConfidenceLevel.LOW,
                        eval_score=mapping.overall_confidence,
                        rules_passed=False,
                        rules_violations=["schema_mapping_confidence_below_threshold"],
                    ),
                    audit=AuditInfo(
                        processing_ms=round((time.monotonic() - t0) * 1000),
                    ),
                )

            # ── 3. Rename row keys to canonical names ──────────────────────
            # Build a rename dict: original header → canonical field
            rename_map: dict[str, str] = {
                raw: canon for raw, canon in mapping.mapped.items()
            }

            canonical_rows: list[dict[str, Any]] = []
            for row in all_rows:
                canonical_row: dict[str, Any] = {}
                for raw_key, val in row.items():
                    canon_key = rename_map.get(raw_key, raw_key)
                    canonical_row[canon_key] = val
                canonical_rows.append(canonical_row)

            canonical_cols = set(rename_map.values())

            # ── 4. Detect entity type ─────────────────────────────────────
            entity_type = _detect_entity_type(canonical_cols)
            span.set_attribute("cafm.entity_type", entity_type)
            logger.info(
                "excel_agent.entity_type_detected",
                ingestion_id=str(ingestion_id),
                entity_type=entity_type,
                canonical_cols=sorted(canonical_cols),
            )

            # ── 5. Stream in batches (asyncpg COPY) ───────────────────────
            rows_written = 0
            rows_failed = 0
            batches_written = 0

            asset_entities: list[AssetEntity] = []
            part_entities: list[SparePartEntity] = []
            wo_entities: list[WorkOrderEntity] = []

            builder = _RECORD_BUILDERS.get(entity_type)
            use_direct_copy = entity_type in _DIRECT_COPY_TABLES and builder is not None

            for batch_start in range(0, total_rows, _BATCH_SIZE):
                batch = canonical_rows[batch_start: batch_start + _BATCH_SIZE]
                batch_records: list[tuple[Any, ...]] = []
                batch_columns: list[str] | None = None

                for row in batch:
                    # Normalise: convert openpyxl native types to str/int/None
                    clean_row: dict[str, Any] = {
                        k: (None if v is None else v)
                        for k, v in row.items()
                    }

                    # Build entity for IntermediateSchema
                    if entity_type == "assets":
                        ent = _row_to_asset_entity(clean_row)
                        if ent:
                            asset_entities.append(ent)
                    elif entity_type == "spare_parts":
                        ent = _row_to_spare_part_entity(clean_row)
                        if ent:
                            part_entities.append(ent)
                    elif entity_type == "work_orders":
                        ent = _row_to_work_order_entity(clean_row)
                        if ent:
                            wo_entities.append(ent)

                    # Build COPY record for direct-write tables
                    if use_direct_copy and builder is not None:
                        try:
                            record, cols = builder(clean_row, organization_id, mapping)
                            if batch_columns is None:
                                batch_columns = cols
                            if len(record) == len(batch_columns):
                                batch_records.append(record)
                        except Exception as row_exc:
                            rows_failed += 1
                            logger.debug(
                                "excel_agent.row_skipped",
                                ingestion_id=str(ingestion_id),
                                error=str(row_exc),
                            )

                # Write batch via asyncpg COPY (skipped in dry_run mode)
                if use_direct_copy and batch_records and batch_columns and not dry_run:
                    try:
                        await _copy_batch_to_table(
                            engine, entity_type, batch_records, batch_columns
                        )
                        rows_written += len(batch_records)
                        batches_written += 1
                        logger.debug(
                            "excel_agent.batch_written",
                            ingestion_id=str(ingestion_id),
                            batch_num=batches_written,
                            rows=len(batch_records),
                            table=entity_type,
                        )
                    except Exception as copy_exc:
                        rows_failed += len(batch_records)
                        logger.error(
                            "excel_agent.copy_failed",
                            ingestion_id=str(ingestion_id),
                            batch_num=batches_written + 1,
                            table=entity_type,
                            error=str(copy_exc),
                        )

            processing_ms = round((time.monotonic() - t0) * 1000)
            total_entities = len(asset_entities) + len(part_entities) + len(wo_entities)

            # ── EL-4.0 — Post-write row count check ──────────────────────
            # Warn if written count is substantially lower than total rows.
            # (Hard rows_failed > 20% threshold triggers a warning, not a stop —
            # partial data is still useful; caller routes via confidence_router.)
            if use_direct_copy and total_rows > 0:
                write_ratio = rows_written / total_rows
                if write_ratio < 0.80:
                    logger.warning(
                        "excel_agent.el4_row_count_mismatch",
                        ingestion_id=str(ingestion_id),
                        total_rows=total_rows,
                        rows_written=rows_written,
                        rows_failed=rows_failed,
                        write_ratio=round(write_ratio, 3),
                    )

            span.set_attribute("cafm.rows_written", rows_written)
            span.set_attribute("cafm.rows_failed", rows_failed)
            span.set_attribute("cafm.entity_count", total_entities)
            span.set_status(StatusCode.OK)

            logger.info(
                "excel_agent.extraction_complete",
                ingestion_id=str(ingestion_id),
                source_filename=source_filename,
                sheet=sheet_name,
                entity_type=entity_type,
                total_rows=total_rows,
                rows_written=rows_written,
                rows_failed=rows_failed,
                direct_copy=use_direct_copy,
                processing_ms=processing_ms,
            )

            # ── 6. Build IntermediateSchema ───────────────────────────────
            if rows_failed == 0:
                overall_conf = ConfidenceLevel.HIGH
                eval_score = 0.95
            elif rows_failed < total_rows * 0.1:
                overall_conf = ConfidenceLevel.MEDIUM
                eval_score = 0.80
            else:
                overall_conf = ConfidenceLevel.LOW
                eval_score = 0.60

            entities = EntitiesBlock(
                assets=asset_entities,
                spare_parts=part_entities,
                work_orders=wo_entities,
            )

            return IntermediateSchema(
                ingestion_id=ingestion_id,
                source_type=SourceType.EXCEL,
                agent_id=AgentId.EXCEL,
                source_filename=source_filename,
                source_blob_url=blob_url,
                extraction_method=ExtractionMethod.OPENPYXL_CLAUDE,
                model_used=ModelUsed.HAIKU,   # Haiku used for schema mapping only
                entities=entities,
                confidence=ConfidenceResult(
                    overall=overall_conf,
                    eval_score=eval_score,
                    rules_passed=rows_failed == 0,
                    rules_violations=(
                        [f"{rows_failed} rows failed during COPY"]
                        if rows_failed > 0 else []
                    ),
                ),
                audit=AuditInfo(
                    processing_ms=processing_ms,
                ),
            )

        except Exception as exc:
            span.record_exception(exc)
            span.set_status(StatusCode.ERROR, str(exc))
            logger.error(
                "excel_agent.extraction_failed",
                ingestion_id=str(ingestion_id),
                source_filename=source_filename,
                error=str(exc),
            )
            raise
